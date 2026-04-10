import requests
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from config import API_KEY

def search_shows(show, broad=False): # Performs search for a given title
    try:
        if broad: # Block only runs if user triggers a broader search for title
            search = requests.get(
                "https://www.omdbapi.com/",
                params = {"apikey": API_KEY, "s": show, "type": "series"}
            ).json()

            if len(search["Search"]) == 1:
                print(f'No other version of "{show}" found.')
                return [search["Search"][0]["imdbID"]] # Returns automatically if only one result is found
            
            print(f'Search results for "{show}":')
            i = 1
            for show in search["Search"]:
                print(f"{i}. {show["Title"]} ({show["Year"]})")
                i += 1
            
            while True:
                selection = input("\nSelect the number(s) of the version(s) you want to fetch, separated by commas (or press Enter for the first result): ")
                if not selection:
                    return [search["Search"][0]["imdbID"]] # First result is returned if user skips
                
                try:
                    selections = [int(s.strip()) for s in selection.split(",")]
                    if any(number < 1 or number > len(search["Search"]) for number in selections):
                        print(f"\nOne or more numbers is invalid. Please select between 1 and {len(selections)}: ")
                        continue
                    else:
                        return [search["Search"][i - 1]["imdbID"] for i in selections] # Or returns a list of imdb IDs based on user's selection
                    
                except ValueError:
                    print("\nPlease enter a number. ")

        search = requests.get( # Finds and returns the first/most popular title the API can find
            "https://www.omdbapi.com/",
            params = {"apikey": API_KEY, "t": show, "type": "series"}
        ).json()

        return search["imdbID"]

    except (KeyError, requests.exceptions.RequestException):
        return None

def fetch_metadata(imdb_id): # Fetches a show's metadata based on its imdb ID
    try:
        scrape = requests.get( 
            "https://www.omdbapi.com/",
            params = {"apikey": API_KEY, "i": imdb_id}
        ).json()

        seasons = scrape["totalSeasons"]

        return {
            "title": scrape["Title"],
            "year": scrape["Year"],
            "seasons": int(seasons) if seasons != "N/A" else 0,
            "imdb_id": imdb_id # Stored for use in get_episodes
        }
    except (KeyError, requests.exceptions.RequestException):
        return None
    
def add_if_unique(list, show):
    if any(s["imdb_id"] == show["imdb_id"] for s in list):
        print(f"\n{show['title']} is already in the confirmed show list.")
        return False
    list.append(show)
    return True
    
def confirm_shows(show_list): # Handles confirmation of show list before episodes are fetched
    confirmed = []
    flagged = [] 

    for show in show_list:
        imdb_id = search_shows(show)
        if imdb_id:
            success = fetch_metadata(imdb_id)
            if success:
                add_if_unique(confirmed, success)
        else:
            flagged.append(show)

    if not confirmed and flagged:
        print("\nNo results found for any of the provided shows.")

    if confirmed:
        print("\nFound the following show(s):\n")
        i = 1
        for show in confirmed:
            print(f"{i}. {show['title']} ({show['year']}, {show['seasons']} seasons)")
            i += 1
            
    if flagged:
        print("\nThe following show(s) could not be found:\n")
        i = 1
        for show in flagged:
            print(f"{i}. {show}")
            i += 1
    
        while True:
            retry = input("\nSearch again for the above show(s)? (y/n): ")
            if retry.lower() in ["y", "n"]:
                break
            print("\nInvalid input. Please enter y or n.")

        if retry.lower() == "y": # If user enters 'n', this block is skipped
            while flagged:
                research_show = input("\nEnter corrected title, starting at #1 (or simply press Enter to skip): ")
                if not research_show:
                    if flagged:
                        print(f"\n{len(flagged)} show(s) skipped.")
                    break
                imdb_id = search_shows(research_show)
                if imdb_id:
                    flagged.pop(0)
                    research_success = fetch_metadata(imdb_id)
                    if research_success:
                        add_if_unique(confirmed, research_success)
                        print(f"\nAdded: {research_success['title']}")

                if flagged:
                    print("\nShows still not found:\n")
                    i = 1
                    for title in flagged:
                        print(f"{i}. {title}")
                        i += 1

        print("\nShows to fetch:")
        i = 1
        for show in confirmed:
            print(f"{i}. {show['title']} ({show['year']}, {show['seasons']} seasons)")
            i += 1

    user_confirm = input("\nConfirm all? (y) Add a show? (a) Change a show? (c) Find a different version of a show? (v) Remove a show? (r) End search? (x): ")
    if user_confirm.lower() == "y":
        return confirmed
    else:
        while user_confirm.lower() != "y":
            # Add block
            if user_confirm.lower() == "a":
                add_show = input('\nEnter the show title to add, or press Enter to skip: ')
                if not add_show:
                    pass
                else:
                    imdb_id = search_shows(add_show)
                    if imdb_id:
                        add_success = fetch_metadata(imdb_id)
                        if add_success:
                            add_if_unique(confirmed, add_success)  
                            print(f"\nAdded: {add_success['title']}")   
            # Change block
            elif user_confirm.lower() == "c":
                change_input = input(f"\nEnter the number and show change (e.g., #. New Show Title), or press Enter to skip: ")
                if not change_input:
                    pass
                else:
                    try:
                        change_number, change_show = change_input.split(". ", 1)
                        change_number = int(change_number)
                        if change_number < 1 or change_number > len(confirmed):
                            print(f"\nInvalid number. Please enter a number between 1 and {len(confirmed)}.")
                            continue
                    except (ValueError, IndexError):
                        print('\nInvalid format. Please enter as "#. New Show Title"')
                        continue
                    imdb_id = search_shows(change_show)
                    if imdb_id:
                        change_success = fetch_metadata(imdb_id)
                        if change_success:
                            if any(s["imdb_id"] == change_success["imdb_id"] for s in confirmed):
                                print(f"\n{change_success['title']} is already in the list. Try a different show.")
                            else:
                                confirmed[change_number - 1] = change_success
            # Different version block
            elif user_confirm.lower() == "v":
                version_input = input(f"\nEnter the show number to search for again, or press Enter to skip: ")
                if not version_input:
                    pass
                else:
                    try:
                        version_number = int(version_input)
                        if version_number < 1 or version_number > len(confirmed):
                            print(f"\nInvalid number. Please enter a number between 1 and {len(confirmed)}.")
                        else:
                            show_version = confirmed[version_number - 1]["title"]
                            imdb_ids = search_shows(show_version, broad=True)
                            if imdb_ids:
                                confirmed.pop(version_number - 1)
                                for imdb_id in imdb_ids:
                                    version_success = fetch_metadata(imdb_id)
                                    if version_success:
                                        add_if_unique(confirmed, version_success)
                    except ValueError:
                        print(f"\nInvalid input. Please enter a number between 1 and {len(confirmed)}.")
            # Remove block
            elif user_confirm.lower() == "r":
                remove_input = input(f"\nEnter the show number to remove, or press Enter to skip: ")
                if not remove_input:
                    pass
                else:
                    try:
                        remove_number = int(remove_input)
                        if remove_number < 1 or remove_number > len(confirmed):
                            print(f"\nInvalid number. Please enter a number between 1 and {len(confirmed)}.")
                        else:
                            del confirmed[remove_number - 1]
                    except ValueError:
                        print(f"\nInvalid input. Please enter a number between 1 and {len(confirmed)}.")
            # Exit block
            elif user_confirm.lower() == "x":
                return []
            else:
                print("\nInvalid input. Please enter y, a, c, v, r, or x.")

            print("\nCurrent show list:\n") # Gives user current state of show list after each alteration made
            i = 1
            for show in confirmed:
                print(f"{i}. {show['title']} ({show['year']}, {show['seasons']} seasons)")
                i += 1

            user_confirm = input("\nConfirm all? (y) Add a show? (a) Change a show? (c) Find a different version of a show? (v) Remove a show? (r) End search? (x): ")
        
    return confirmed

def fill_missing_dates(episodes): # Interpolates any "N/A" airdates
    for current in episodes:
        if current["Released"] == "N/A":
            # Find all predecessor and successor dated episodes
            predecessors = [ep for ep in episodes if ep["Released"] != "N/A" and int(ep["Episode"]) < int(current["Episode"])]
            successors = [ep for ep in episodes if ep["Released"] != "N/A" and int(ep["Episode"]) > int(current["Episode"])]
            
            # Find closest dated episode, prioritizing predecessors
            if predecessors:
                anchor = max(predecessors, key=lambda ep: int(ep["Episode"]))
            elif successors:
                anchor = min(successors, key=lambda ep: int(ep["Episode"]))
            else:
                anchor = None
            
            # Estimate date based on episode's distance from anchor
            if anchor:                
                distance = abs(int(current["Episode"]) - int(anchor["Episode"]))
                delta = timedelta(days=7 * distance)
                date = datetime.strptime(anchor["Released"], "%Y-%m-%d")
                if int(anchor["Episode"]) < int(current["Episode"]):
                    current["Released"] = (date + delta).strftime("%Y-%m-%d")
                else:
                    current["Released"] = (date - delta).strftime("%Y-%m-%d")
                current["Note"] = "Airdate is an estimate."
            else:
                current["Released"] = "9999-99-99"
                current["Note"] = "Airdate missing."

    return episodes

def get_episodes(confirmed): # Fetches show's episodes
    all_episodes = []

    for show in confirmed:
        for season in range(1, show["seasons"] + 1):
            season_data = requests.get( # Fetch episode list for each season
                "https://www.omdbapi.com/",
                params={"apikey": API_KEY, "i": show["imdb_id"], "Season": season}
            ).json()

            episodes = season_data["Episodes"]
            episodes = fill_missing_dates(episodes) # Fixes any "N/A" airdate issues
            seen = {}
            for episode in episodes:
                key = episode["Episode"]
                if key not in seen or not re.match(r"^Episode #\d+\.\d+$", episode["Title"]):
                    seen[key] = episode

            # Add show title and season # to each ep, then adds the finalized ep to the cleaned list (errant eps are filtered) 
            for episode in seen.values():
                if re.match(r"^Episode #\d+\.\d+$", episode["Title"]):
                    continue
                episode["Show"] = show["title"]
                episode["Season"] = season
                all_episodes.append(episode)

    return all_episodes

def format_date(date): # Converts yyyy-mm-dd to 'Month dd, yyyy' for output
    if date == "9999-99-99":
        return "Unknown"
    return datetime.strptime(date, "%Y-%m-%d").strftime("%B %d, %Y")

def write_file(watch_list, file_name):
    # .xlsx file
    wb = Workbook()
    ws = wb.active
    ws.title = f"{file_name} Watch List"

    has_notes = any(episode.get("Note") for episode in watch_list)

    # Header row
    ws.append(["Show", "Season & Episode", "Title", "Airdate"] + (["Note"] if has_notes else []))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    colors =["E69F00", "56B4E9", "009E73", "F0E442", "0072B2", "D55E00", "CC79A7", "999999", "FFFFFF"]
    light_colors = ["F0E442", "999999", "FFFFFF"]
    show_titles = list(dict.fromkeys(ep["Show"] for ep in watch_list))
    show_colors = {show: colors[i % len(colors)] for i, show in enumerate(show_titles)}

    for episode in watch_list:
        ws.append([
        episode['Show'],
        f"S{episode['Season']}E{episode['Episode']}",
        episode['Title'],
        format_date(episode['Released'])]
        + ([episode.get('Note', '')] if has_notes else []))

        fill = PatternFill(start_color=show_colors[episode['Show']], end_color=show_colors[episode['Show']], fill_type="solid")
        font_color = "000000" if show_colors[episode['Show']] in light_colors else "FFFFFF"
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = Font(color=font_color)
    
    # Filters
    ws.auto_filter.ref = ws.dimensions

    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    wb.save(f"{file_name}.xlsx")

universe = input('Enter shows, separated by commas: ')
uni_list = [show.strip() for show in universe.split(",")] # Parses each user-input show

confirmed = confirm_shows(uni_list)

if not confirmed:
    print("\nSearch ended.")
else:
    watch_list = get_episodes(confirmed)
    watch_list.sort(key=lambda ep: ep["Released"]) # Sorts final list by airdate
    file_name = input("\nNow, provide a name for the watch list: ")
    write_file(watch_list, file_name)